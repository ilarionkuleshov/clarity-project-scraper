import os
import openpyxl

from twisted.internet import reactor, defer
from sqlalchemy.dialects.mysql import insert

from .base import DatabaseReactorCommand
from database.models import Finances
from utils import compile_and_stringify_statement


class XlsxEdrImporter(DatabaseReactorCommand):
    chunk_size = 100
    edr_counter = 0

    def add_options(self, parser):
        super().add_options(parser)
        parser.add_option(
            "-f",
            "--file",
            default=None,
            dest="xlsx_file",
            help="xlsx file for edr's import to DB"
        )

    def execute(self, args, opts):
        if opts.xlsx_file and os.path.isfile(opts.xlsx_file):
            self.logger.info("Reading xlsx file...")
            wookbook = openpyxl.load_workbook(opts.xlsx_file)
            worksheet = wookbook.active

            edr_values = []
            deferred_interactions = []
            for row in worksheet.iter_rows(2, worksheet.max_row):
                self.edr_counter += 1
                edr_values.append(
                    {"edr": self.validate_edr(row[0].value)}
                )
                if self.edr_counter % self.chunk_size == 0:
                    d = self.db_connection_pool.runQuery(
                        compile_and_stringify_statement(
                            insert(Finances).values(edr_values).prefix_with("IGNORE")
                        )
                    )
                    d.addCallback(self.on_edr_insert_success)
                    deferred_interactions.append(d)
                    edr_values = []
            deferred_list = defer.DeferredList(deferred_interactions, consumeErrors=True)
            deferred_list.addCallback(self.on_total_success)
            deferred_list.addErrback(self.errback)
            return deferred_list
        else:
            self.logger.error("Argument `--file` is incorrect")
            deferred = defer.Deferred()
            reactor.callLater(0, deferred.callback, True)
            return deferred

    def on_edr_insert_success(self, _):
        self.logger.info(f"Inserted {self.chunk_size} edr`s")

    def on_total_success(self, _):
        self.logger.info(f"Successfully inserted {self.edr_counter} edr`s")

    @staticmethod
    def validate_edr(edr):
        edr = edr.strip()
        return edr if edr.isdigit() else None
